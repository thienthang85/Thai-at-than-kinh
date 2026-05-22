import pandas as pd
import datetime
from datetime import timedelta
from borax.calendars.lunardate import LunarDate, TermUtils
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

class ThaiAtStarsMap:
    CUNG_BAN = ["Tý", "Sửu", "Dần", "Mão", "Thìn", "Tỵ", "Ngọ", "Mùi", "Thân", "Dậu", "Tuất", "Hợi"]
    CUU_CUNG_TO_12_CHI = {
        1: ["Tuất", "Hợi"], 2: ["Ngọ"], 3: ["Sửu", "Dần"], 4: ["Mão"], 5: ["Trung Cung"],
        6: ["Dậu"], 7: ["Mùi", "Thân"], 8: ["Tý"], 9: ["Thìn", "Tỵ"]
    }
    
    STAR_RANKS = {
        # Sinh Khí
        "Ngũ Phúc": {"Trên": ["Thìn", "Sửu", "Thân", "Hợi"], "Trung": ["Tuất", "Ngọ", "Tỵ", "Dậu", "Mùi", "Tý"], "Dưới": ["Dần", "Mão"]},
        "Quân Cơ": {"Trên": ["Thìn", "Ngọ", "Tuất", "Sửu", "Mùi"], "Trung": ["Tỵ", "Thân", "Hợi", "Tý"], "Dưới": ["Dần", "Mão", "Dậu"]},
        "Thần Cơ": {"Trên": ["Thìn", "Tuất", "Sửu", "Mão"], "Trung": ["Mùi", "Thân", "Ngọ", "Hợi"], "Dưới": ["Dậu", "Tỵ", "Tý", "Dần"]},
        "Dân Cơ": {"Trên": ["Thân", "Thìn", "Hợi", "Tý"], "Trung": ["Mùi", "Tỵ", "Ngọ", "Tuất", "Sửu"], "Dưới": ["Dần", "Mão", "Dậu"]},
        "Văn Xương": {"Trên": ["Thân", "Sửu", "Thìn", "Hợi"], "Trung": ["Tỵ", "Ngọ", "Mùi", "Dậu", "Tuất"], "Dưới": ["Dần", "Mão", "Tý"]},
        "Thần Kể": {"Trên": ["Thìn", "Tuất", "Sửu", "Mùi"], "Trung": ["Thân", "Tý", "Hợi", "Tỵ", "Ngọ", "Dậu"], "Dưới": ["Dần", "Mão"]},
        
        # Tướng Soái
        "Thái Ất": {"Trên": ["Dần", "Hợi", "Mão", "Mùi"], "Trung": ["Sửu", "Thìn", "Tý", "Tuất"], "Dưới": ["Thân", "Dậu", "Tỵ", "Ngọ"]},
        "Đại Tướng Chủ": {"Trên": ["Tỵ", "Dậu", "Sửu", "Thân"], "Trung": ["Tý", "Mão", "Thìn", "Mùi", "Hợi"], "Dưới": ["Dần", "Ngọ", "Tuất"]},
        "Đại Tướng Khách": {"Trên": ["Thân", "Tý", "Thìn", "Hợi"], "Trung": ["Mùi", "Sửu", "Dần", "Dậu"], "Dưới": ["Tỵ", "Ngọ", "Mão", "Tuất"]},
        "Tham Tướng Chủ": {"Trên": ["Thân", "Tý", "Thìn", "Hợi"], "Trung": ["Tỵ", "Dậu", "Ngọ", "Dần"], "Dưới": ["Sửu", "Mùi", "Tuất", "Mão"]},
        "Tham Tướng Khách": {"Trên": ["Hợi", "Mão", "Mùi", "Dần"], "Trung": ["Thân", "Tý", "Thìn", "Sửu"], "Dưới": ["Tỵ", "Dậu", "Ngọ", "Tuất"]},
        
        # Hung Họa
        "Thủy Kích": {"Trên": ["Dần", "Ngọ", "Tuất", "Tỵ"], "Trung": ["Thìn", "Mùi"], "Dưới": ["Sửu", "Hợi", "Tý", "Mão", "Dậu", "Thân"]},
        "Tứ Thần": {"Trên": ["Thân", "Tý", "Thìn", "Hợi"], "Trung": ["Ngọ", "Sửu", "Tuất"], "Dưới": ["Mùi", "Mão", "Dần", "Dậu", "Tỵ"]},
        "Thiên Ất": {"Trên": ["Dậu", "Sửu", "Thân"], "Trung": ["Tý", "Thìn", "Mão", "Mùi", "Tuất", "Hợi"], "Dưới": ["Tỵ", "Ngọ", "Dần"]},
        "Địa Ất": {"Trên": ["Thìn", "Tuất", "Sửu", "Mùi"], "Trung": ["Tỵ", "Ngọ", "Thân", "Tý"], "Dưới": ["Dần", "Mão", "Dậu", "Hợi"]},
        "Phi Phù": {"Trên": ["Dần", "Ngọ", "Tuất", "Tỵ"], "Trung": ["Sửu", "Mùi", "Thìn", "Thân"], "Dưới": ["Hợi", "Tý", "Mão", "Dậu"]},
        "Trực Phù": {"Trên": ["Dần", "Ngọ", "Tuất", "Tỵ"], "Trung": ["Tý", "Mão", "Dậu", "Hợi"], "Dưới": ["Sửu", "Mùi", "Thìn", "Thân"]}
    }

    @staticmethod
    def get_star_rank(star_name, location):
        if location in [1, 3, 7, 9]:
            # Rơi vào 4 cung góc thì không xét bậc, ngoại trừ Ngũ Phúc ở Khôn (7) -> Trên
            if star_name == "Ngũ Phúc" and location == 7:
                return f"{star_name} (Trên)"
            return star_name
            
        cung_to_chi = {2: "Ngọ", 4: "Mão", 6: "Dậu", 8: "Tý"}
        chi = location
        if isinstance(location, int):
            chi = cung_to_chi.get(location)
            
        # Thay Vị bằng Mùi để map chuẩn
        if chi == "Vị": chi = "Mùi"
            
        if not chi: return star_name
        
        ranks = ThaiAtStarsMap.STAR_RANKS.get(star_name)
        if not ranks: return star_name
        
        for rank_name, chi_list in ranks.items():
            if chi in chi_list:
                return f"{star_name} ({rank_name})"
                
        return star_name
    GRID_MAPPING = {
        "Tốn": (0, 0), "Tỵ": (0, 1), "Ngọ": (0, 2), "Mùi": (0, 3), "Khôn": (0, 4),
        "Thân": (1, 4), "Dậu": (2, 4), "Tuất": (3, 4), "Kiền": (4, 4),
        "Hợi": (4, 3), "Tý": (4, 2), "Sửu": (4, 1), "Cấn": (4, 0),
        "Dần": (3, 0), "Mão": (2, 0), "Thìn": (1, 0),
        "Trung Cung": (2, 2)
    }
    CUNG_GOC_TO_NAME = {1: "Kiền", 3: "Cấn", 7: "Khôn", 9: "Tốn"}

    @staticmethod
    def calculate_cung_star(tich_nhat, chu_ky, toan_per_cung, khoi_diem, chieu, is_9_cung=True, doanh_sai=0):
        cungs = [1,2,3,4,5,6,7,8,9] if is_9_cung else [1,2,3,4,6,7,8,9]
        total_toan = tich_nhat + doanh_sai
        du = total_toan % chu_ky
        if du == 0: du = chu_ky
        idx = (du - 1) // toan_per_cung
        if idx >= len(cungs): idx = len(cungs) - 1
        start_idx = cungs.index(khoi_diem)
        final_idx = (start_idx + idx) % len(cungs) if chieu == 1 else (start_idx - idx) % len(cungs)
        return cungs[final_idx]

    @staticmethod
    def calculate_chi_star(tich_nhat, chu_ky, toan_per_chi, khoi_diem_chi, chieu, doanh_sai=0):
        chis = ThaiAtStarsMap.CUNG_BAN
        total_toan = tich_nhat + doanh_sai
        du = total_toan % chu_ky
        if du == 0: du = chu_ky
        idx = (du - 1) // toan_per_chi
        idx = idx % len(chis)
        start_idx = chis.index(khoi_diem_chi)
        final_idx = (start_idx + idx) % len(chis) if chieu == 1 else (start_idx - idx) % len(chis)
        return chis[final_idx]

    @staticmethod
    def calculate_ngu_phuc(tich_nhat, doanh_sai=115):
        path = [1, 3, 9, 7, 5]  # Kiền, Cấn, Tốn, Khôn, Trung Cung
        total = tich_nhat + doanh_sai
        du = total % 225
        if du == 0: du = 225
        idx = (du - 1) // 45
        return path[idx]

    @staticmethod
    def calculate_van_xuong_name(tich_nhat, khoi_duong=True):
        if khoi_duong:
            path = ["Thân", "Dậu", "Tuất", "Kiền", "Kiền", "Hợi", "Tý", "Sửu", "Cấn", "Dần", "Mão", "Thìn", "Tốn", "Tỵ", "Ngọ", "Mùi", "Khôn", "Khôn"]
        else:
            path = ["Dần", "Mão", "Thìn", "Tốn", "Tốn", "Tỵ", "Ngọ", "Mùi", "Khôn", "Thân", "Dậu", "Tuất", "Kiền", "Hợi", "Tý", "Sửu", "Cấn", "Cấn"]
        return path[(tich_nhat - 1) % 18]

    @staticmethod
    def calculate_van_xuong(tich_nhat, khoi_duong=True):
        val = ThaiAtStarsMap.calculate_van_xuong_name(tich_nhat, khoi_duong)
        name_to_cung = {
            "Tuất": 1, "Hợi": 1, "Kiền": 1,
            "Ngọ": 2,
            "Sửu": 3, "Dần": 3, "Cấn": 3,
            "Mão": 4,
            "Trung Cung": 5,
            "Dậu": 6,
            "Mùi": 7, "Thân": 7, "Khôn": 7,
            "Tý": 8,
            "Thìn": 9, "Tỵ": 9, "Tốn": 9
        }
        return name_to_cung.get(val, 5)

    @staticmethod
    def map_stars(tich_nhat, khoi_duong=True, bat_tu=None):
        thai_at_cung = ThaiAtStarsMap.calculate_cung_star(tich_nhat, 24, 3, 1 if khoi_duong else 9, 1 if khoi_duong else -1, is_9_cung=False)
        van_xuong_name = ThaiAtStarsMap.calculate_van_xuong_name(tich_nhat, khoi_duong)
        van_xuong_cung = ThaiAtStarsMap.calculate_van_xuong(tich_nhat, khoi_duong)
        thuy_kich_cung = ThaiAtStarsMap.calculate_cung_star(tich_nhat, 9, 1, 3, -1, is_9_cung=True)
        ngu_phuc_cung = ThaiAtStarsMap.calculate_ngu_phuc(tich_nhat)

        cung_to_name = {1: "Kiền", 2: "Ngọ", 3: "Cấn", 4: "Mão", 6: "Dậu", 7: "Khôn", 8: "Tý", 9: "Tốn"}
        
        # Thần Kể: Khối Dương khởi Dần chuyển NGƯỢC 12 thần (Dần,Sửu,Tý,Hợi,Tuất,Dậu,Thân,Mùi,Ngọ,Tỵ,Thìn,Mão)
        chis = ["Tý", "Sửu", "Dần", "Mão", "Thìn", "Tỵ", "Ngọ", "Mùi", "Thân", "Dậu", "Tuất", "Hợi"]
        du_ke = tich_nhat % 12
        if du_ke == 0: du_ke = 12
        start_ke = chis.index("Dần")
        than_ke_name = chis[(start_ke - (du_ke - 1)) % 12]
        
        # Thủy Kích Kể Ngày: Khoảng cách từ Thần Kể đến Văn Xương, rồi khởi Cấn đi thuận
        circle = ["Tý", "Sửu", "Cấn", "Dần", "Mão", "Thìn", "Tốn", "Tỵ", "Ngọ", "Mùi", "Khôn", "Thân", "Dậu", "Tuất", "Kiền", "Hợi"]
        idx_ke = circle.index(than_ke_name)
        idx_vx = circle.index(van_xuong_name)
        steps_tk = (idx_vx - idx_ke) % 16 + 1
        idx_can = circle.index("Cấn")
        thuy_kich_name = circle[(idx_can + steps_tk - 1) % 16]
        
        than_hop = {"Tý": "Sửu", "Sửu": "Tý", "Dần": "Hợi", "Hợi": "Dần", "Mão": "Tuất", "Tuất": "Mão", "Thìn": "Dậu", "Dậu": "Thìn", "Tỵ": "Thân", "Thân": "Tỵ", "Ngọ": "Mùi", "Mùi": "Ngọ"}
        
        # Thái Tuế Kể Ngày là Chi Ngày
        thai_tue = bat_tu["chi_ngay"] if bat_tu else "Sửu"
        hop_chi = than_hop.get(thai_tue, "Tý")
        
        # Kể Định Kể Ngày: Từ Thần Hợp tới Văn Xương, khởi Thái Tuế đếm thuận
        idx_hop = circle.index(hop_chi)
        steps_kd = (idx_vx - idx_hop) % 16 + 1
        idx_tue = circle.index(thai_tue)
        ke_dinh_name = circle[(idx_tue + steps_kd - 1) % 16]

        # --- PHI PHÙ: Tra Bảng Lập Thành Khối Dương (Cục + position) ---
        # Trích từ Cuốn 7 Bảng II - Bảng Phi Phù Lập Thành (chính thức)
        # "Sao Phi Phù an tại cung nào tùy theo cục" (P23573)
        PHI_PHU_KHOI_DUONG = {
            1: "Tốn", 2: "Tốn", 3: "Tốn", 4: "Tốn", 5: "Tốn", 6: "Tốn",
            7: "Khôn", 8: "Khôn", 9: "Khôn", 10: "Kiền", 11: "Kiền", 12: "Kiền",
            13: "Tốn", 14: "Tốn", 15: "Tốn", 16: "Khôn", 17: "Khôn", 18: "Khôn",
            19: "Kiền", 20: "Kiền", 21: "Kiền", 22: "Dậu", 23: "Dậu", 24: "Dậu",
            25: "Trung", 26: "Trung", 27: "Trung", 28: "Dậu", 29: "Dậu", 30: "Dậu",
            31: "Khôn", 32: "Khôn", 33: "Khôn", 34: "Tý", 35: "Tý", 36: "Tý",
            37: "Tốn", 38: "Tốn", 39: "Tốn", 40: "Tốn", 41: "Tốn", 42: "Tốn",
            43: "Khôn", 44: "Khôn", 45: "Khôn", 46: "Kiền", 47: "Kiền", 48: "Kiền",
            49: "Kiền", 50: "Kiền", 51: "Kiền", 52: "Ngọ", 53: "Ngọ", 54: "Ngọ",
            55: "Cấn", 56: "Cấn", 57: "Cấn", 58: "Mão", 59: "Mão", 60: "Mão",
            61: "Trung", 62: "Trung", 63: "Trung", 64: "Dậu", 65: "Dậu", 66: "Dậu",
            67: "Khôn", 68: "Khôn", 69: "Khôn", 70: "Tý", 71: "Tý", 72: "Tý",
        }
        nguyen_cuc = ThaiAtCalculator().calculate_nguyen_cuc(tich_nhat)
        cuc_num = nguyen_cuc.get("cuc", 1)
        phi_phu_pos = PHI_PHU_KHOI_DUONG.get(cuc_num, "Kiền")
        
        # Chuyển tên cung thành số cung nếu là Bát Quái (bao gồm Trung = Cung 5)
        name_to_cung = {"Kiền": 1, "Ngọ": 2, "Cấn": 3, "Mão": 4, "Trung": 5, "Dậu": 6, "Khôn": 7, "Tý": 8, "Tốn": 9}
        
        stars_cung = {
            ThaiAtStarsMap.get_star_rank("Thái Ất", thai_at_cung): thai_at_cung,
            ThaiAtStarsMap.get_star_rank("Ngũ Phúc", ngu_phuc_cung): ngu_phuc_cung,
            "Mã 3": 9, 
            "Phi Lộc": 6, "Mã 8": 8, 
        }
        
        # Phi Phù: đưa vào cung hoặc chi tùy vị trí
        if phi_phu_pos in name_to_cung:
            stars_cung[ThaiAtStarsMap.get_star_rank("Phi Phù", name_to_cung[phi_phu_pos])] = name_to_cung[phi_phu_pos]
        else:
            pass  # Chi positions handled in stars_chi below
        
        if thuy_kich_name in cung_to_name.values():
            for k, v in cung_to_name.items():
                if v == thuy_kich_name: stars_cung[ThaiAtStarsMap.get_star_rank("Thủy Kích", k)] = k
                
        if ke_dinh_name in cung_to_name.values():
            for k, v in cung_to_name.items():
                if v == ke_dinh_name: stars_cung["Kể Định"] = k
                
        quan_co_chi = ThaiAtStarsMap.calculate_chi_star(tich_nhat, 360, 30, "Ngọ", 1, doanh_sai=250)
        than_co_chi = ThaiAtStarsMap.calculate_chi_star(tich_nhat, 36, 3, "Ngọ", 1, doanh_sai=250)
        dan_co_chi = ThaiAtStarsMap.calculate_chi_star(tich_nhat, 12, 1, "Tuất", 1, doanh_sai=250)
        
        # --- NHÓM 4 SAO ĐI CHÙM: Tứ Thần, Thiên Ất, Địa Ất, Trực Phù ---
        # Chu kỳ 36, 3 ngày/cung, quỹ đạo 12 vị trí thuận
        orbit_12 = [1, 2, 3, 4, 5, 6, 7, 8, 9, "Tỵ", "Thân", "Dần"]
        du_36 = tich_nhat % 36
        if du_36 == 0: du_36 = 36
        offset_4star = (du_36 - 1) // 3  # Số cung đã đi qua
        
        star_starts = {"Tứ Thần": 1, "Trực Phù": 5, "Thiên Ất": 6, "Địa Ất": 9}
        star_4_positions = {}
        for sname, start_cung in star_starts.items():
            start_idx = orbit_12.index(start_cung)
            pos = orbit_12[(start_idx + offset_4star) % 12]
            star_4_positions[sname] = pos
        
        # Chuyển vị trí sang tên Chi (nếu là số cung thì tra cung_to_name)
        def pos_to_chi(pos):
            if isinstance(pos, str): return pos  # Đã là Chi (Tỵ, Thân, Dần)
            return cung_to_name.get(pos, f"Cung {pos}")
        
        tu_than_chi = pos_to_chi(star_4_positions["Tứ Thần"])
        thien_at_chi = pos_to_chi(star_4_positions["Thiên Ất"])
        dia_at_chi = pos_to_chi(star_4_positions["Địa Ất"])
        truc_phu_chi = pos_to_chi(star_4_positions["Trực Phù"])
        
        # --- SAO THẺ ĐẾ (Đế Phù) ---
        # Tích / 20 lấy dư. Khởi Tuất đi thuận (bỏ Kiền), lưu 2 toán tại Tý/Ngọ/Mão/Dậu
        sao_the_path = ["Tuất", "Hợi", "Tý", "Tý", "Sửu", "Cấn", "Dần",
                        "Mão", "Mão", "Thìn", "Tốn", "Tỵ", "Ngọ", "Ngọ",
                        "Mùi", "Khôn", "Thân", "Dậu", "Dậu"]  # 19 vị trí
        du_the = tich_nhat % 20
        if du_the == 0: du_the = 19  # Hết vòng = vị trí cuối
        sao_the_chi = sao_the_path[min(du_the - 1, 18)]
        
        # --- THÁI ÂM (Âm Cả) ---
        # Luôn đi trước Thái Tuế 2 cung
        thai_tue_idx = chis.index(thai_tue)
        thai_am_chi = chis[(thai_tue_idx - 2) % 12]
        
        stars_chi = {
            ThaiAtStarsMap.get_star_rank("Quân Cơ", quan_co_chi): quan_co_chi,
            ThaiAtStarsMap.get_star_rank("Thần Cơ", than_co_chi): than_co_chi,
            ThaiAtStarsMap.get_star_rank("Dân Cơ", dan_co_chi): dan_co_chi,
            ThaiAtStarsMap.get_star_rank("Tứ Thần", tu_than_chi): tu_than_chi,
            ThaiAtStarsMap.get_star_rank("Thiên Ất", thien_at_chi): thien_at_chi,
            ThaiAtStarsMap.get_star_rank("Trực Phù", truc_phu_chi): truc_phu_chi,
            "Thái Âm": thai_am_chi,
            "Sao Thẻ": sao_the_chi,
            "Thái Tuế": thai_tue,
            "Thần Hợp": hop_chi
        }
        
        # Địa Ất: có thể rơi vào cung góc hoặc chi
        dia_at_pos = star_4_positions["Địa Ất"]
        if isinstance(dia_at_pos, int) and dia_at_pos in cung_to_name:
            stars_cung[ThaiAtStarsMap.get_star_rank("Địa Ất", dia_at_pos)] = dia_at_pos
        else:
            stars_chi[ThaiAtStarsMap.get_star_rank("Địa Ất", dia_at_chi)] = dia_at_chi
        
        stars_chi[ThaiAtStarsMap.get_star_rank("Thần Kể", than_ke_name)] = than_ke_name
        
        if thuy_kich_name not in cung_to_name.values():
            stars_chi[ThaiAtStarsMap.get_star_rank("Thủy Kích", thuy_kich_name)] = thuy_kich_name
        
        if van_xuong_name in cung_to_name.values():
            for k, v in cung_to_name.items():
                if v == van_xuong_name: stars_cung[ThaiAtStarsMap.get_star_rank("Văn Xương", k)] = k
        else:
            stars_chi[ThaiAtStarsMap.get_star_rank("Văn Xương", van_xuong_name)] = van_xuong_name
            
        if ke_dinh_name not in cung_to_name.values():
            stars_chi["Kể Định"] = ke_dinh_name
            
        def calc_toan(start_name, end_name):
            c = [("Tý", "Chính", 8), ("Sửu", "Gián", 0), ("Cấn", "Chính", 3), ("Dần", "Gián", 0),
                 ("Mão", "Chính", 4), ("Thìn", "Gián", 0), ("Tốn", "Chính", 9), ("Tỵ", "Gián", 0),
                 ("Ngọ", "Chính", 2), ("Mùi", "Gián", 0), ("Khôn", "Chính", 7), ("Thân", "Gián", 0),
                 ("Dậu", "Chính", 6), ("Tuất", "Gián", 0), ("Kiền", "Chính", 1), ("Hợi", "Gián", 0)]
            names = [x[0] for x in c]
            s = names.index(start_name)
            e = names.index(end_name)
            
            # Ngoại lệ: Trùng cung thì không di chuyển, lấy luôn biệt số
            if s == e:
                return 1 if c[s][1] == "Gián" else c[s][2]
                
            total = 1 if c[s][1] == "Gián" else c[s][2]
            curr = (s + 1) % 16
            while curr != e:
                if c[curr][1] == "Chính": total += c[curr][2]
                curr = (curr + 1) % 16
            return total

        thai_at_name = cung_to_name.get(thai_at_cung, "Tốn")
        
        toan_chu = calc_toan(van_xuong_name, thai_at_name)
        toan_khach = calc_toan(thuy_kich_name, thai_at_name)
        toan_dinh = calc_toan(ke_dinh_name, thai_at_name)
        
        def get_dai_tuong(val):
            u = val % 10
            return val // 10 if u == 0 else u
            
        dai_chu = get_dai_tuong(toan_chu)
        dai_khach = get_dai_tuong(toan_khach)
        tham_chu = get_dai_tuong(dai_chu * 3)
        tham_khach = get_dai_tuong(dai_khach * 3)
        
        stars_cung[ThaiAtStarsMap.get_star_rank("Đại Tướng Chủ", dai_chu)] = dai_chu
        stars_cung[ThaiAtStarsMap.get_star_rank("Đại Tướng Khách", dai_khach)] = dai_khach
        stars_cung[ThaiAtStarsMap.get_star_rank("Tham Tướng Chủ", tham_chu)] = tham_chu
        stars_cung[ThaiAtStarsMap.get_star_rank("Tham Tướng Khách", tham_khach)] = tham_khach
        
        stars_cung["_toan"] = {
            "chu": toan_chu, "khach": toan_khach, "dinh": toan_dinh,
            "van_xuong_pos": van_xuong_name, "thuy_kich_pos": thuy_kich_name,
            "thai_at_cung": thai_at_cung
        }
        
        return stars_cung, stars_chi

class ThaiAtCalculator:
    HEXAGRAMS = {
        1: {"name": "THUẦN KIỀN", "binary": "111111"},
        2: {"name": "THUẦN KHÔN", "binary": "000000"},
        3: {"name": "THỦY LÔI TRUÂN", "binary": "100010"},
        4: {"name": "SƠN THỦY MÔNG", "binary": "010001"},
        5: {"name": "THỦY THIÊN NHU", "binary": "111010"},
        6: {"name": "THIÊN THỦY TỤNG", "binary": "010111"},
        7: {"name": "ĐỊA THỦY SƯ", "binary": "010000"},
        8: {"name": "THỦY ĐỊA TỶ", "binary": "000010"},
        9: {"name": "PHONG THIÊN TIỂU SÚC", "binary": "111011"},
        10: {"name": "THIÊN TRẠCH LÝ", "binary": "110111"},
        11: {"name": "ĐỊA THIÊN THÁI", "binary": "111000"},
        12: {"name": "THIÊN ĐỊA BĨ", "binary": "000111"},
        13: {"name": "THIÊN HỎA ĐỒNG NHÂN", "binary": "101111"},
        14: {"name": "HỎA THIÊN ĐẠI HỮU", "binary": "111101"},
        15: {"name": "ĐỊA SƠN KHIÊM", "binary": "001000"},
        16: {"name": "LÔI ĐỊA DỰ", "binary": "000100"},
        17: {"name": "TRẠCH LÔI TÙY", "binary": "100110"},
        18: {"name": "SƠN PHONG CỔ", "binary": "011001"},
        19: {"name": "ĐỊA TRẠCH LÂM", "binary": "110000"},
        20: {"name": "PHONG ĐỊA QUAN", "binary": "000011"},
        21: {"name": "HỎA LÔI PHỆ HẠP", "binary": "100101"},
        22: {"name": "SƠN HỎA BÍ", "binary": "101001"},
        23: {"name": "SƠN ĐỊA BÁC", "binary": "000001"},
        24: {"name": "ĐỊA LÔI PHỤC", "binary": "100000"},
        25: {"name": "THIÊN LÔI VÔ VỌNG", "binary": "100111"},
        26: {"name": "SƠN THIÊN ĐẠI SÚC", "binary": "111001"},
        27: {"name": "SƠN LÔI DI", "binary": "100001"},
        28: {"name": "TRẠCH PHONG ĐẠI QUÁ", "binary": "011110"},
        29: {"name": "THUẦN KHẢM", "binary": "010010"},
        30: {"name": "THUẦN LY", "binary": "101101"},
        31: {"name": "TRẠCH SƠN HÀM", "binary": "001110"},
        32: {"name": "LÔI PHONG HẰNG", "binary": "011100"},
        33: {"name": "THIÊN SƠN ĐỘN", "binary": "001111"},
        34: {"name": "LÔI THIÊN ĐẠI TRÁNG", "binary": "111100"},
        35: {"name": "HỎA ĐỊA TẤN", "binary": "000101"},
        36: {"name": "ĐỊA HỎA MINH DI", "binary": "101000"},
        37: {"name": "PHONG HỎA GIA NHÂN", "binary": "101011"},
        38: {"name": "HỎA TRẠCH KHUÊ", "binary": "110101"},
        39: {"name": "THỦY SƠN KIỂN", "binary": "001010"},
        40: {"name": "LÔI THỦY GIẢI", "binary": "010100"},
        41: {"name": "SƠN TRẠCH TỔN", "binary": "110001"},
        42: {"name": "PHONG LÔI ÍCH", "binary": "100011"},
        43: {"name": "TRẠCH THIÊN QUẢI", "binary": "111110"},
        44: {"name": "THIÊN PHONG CẤU", "binary": "011111"},
        45: {"name": "TRẠCH ĐỊA TỤY", "binary": "000110"},
        46: {"name": "ĐỊA PHONG THĂNG", "binary": "011000"},
        47: {"name": "TRẠCH THỦY KHỐN", "binary": "010110"},
        48: {"name": "THỦY PHONG TỈNH", "binary": "011010"},
        49: {"name": "TRẠCH HỎA CÁCH", "binary": "101110"},
        50: {"name": "HỎA PHONG ĐỈNH", "binary": "011101"},
        51: {"name": "THUẦN CHẤN", "binary": "100100"},
        52: {"name": "THUẦN CẤN", "binary": "001001"},
        53: {"name": "PHONG SƠN TIỆM", "binary": "001011"},
        54: {"name": "LÔI TRẠCH QUY MUỘI", "binary": "110100"},
        55: {"name": "LÔI HỎA PHONG", "binary": "101100"},
        56: {"name": "HỎA SƠN LỮ", "binary": "001101"},
        57: {"name": "THUẦN TỐN", "binary": "011011"},
        58: {"name": "THUẦN ĐOÀI", "binary": "110110"},
        59: {"name": "PHONG THỦY HOÁN", "binary": "010011"},
        60: {"name": "THỦY TRẠCH TIẾT", "binary": "110010"},
        61: {"name": "PHONG TRẠCH TRUNG PHU", "binary": "110011"},
        62: {"name": "LÔI SƠN TIỂU QUÁ", "binary": "001100"},
        63: {"name": "THỦY HỎA KÝ TẾ", "binary": "101010"},
        64: {"name": "HỎA THỦY VỊ TẾ", "binary": "010101"}
    }

    def __init__(self):
        self.thien_can = ["Giáp", "Ất", "Bính", "Đinh", "Mậu", "Kỷ", "Canh", "Tân", "Nhâm", "Quý"]
        self.dia_chi = ["Tý", "Sửu", "Dần", "Mão", "Thìn", "Tỵ", "Ngọ", "Mùi", "Thân", "Dậu", "Tuất", "Hợi"]

    def calculate_bat_tu(self, dt: datetime.datetime):
        try:
            ld = LunarDate.from_solar_date(dt.year, dt.month, dt.day)
            can_map = {'甲':'Giáp', '乙':'Ất', '丙':'Bính', '丁':'Đinh', '戊':'Mậu', '己':'Kỷ', '庚':'Canh', '辛':'Tân', '壬':'Nhâm', '癸':'Quý'}
            chi_map = {'子':'Tý', '丑':'Sửu', '寅':'Dần', '卯':'Mão', '辰':'Thìn', '巳':'Tỵ', '午':'Ngọ', '未':'Mùi', '申':'Thân', '酉':'Dậu', '戌':'Tuất', '亥':'Hợi'}
            
            def translate_gz(gz_str):
                can = can_map.get(gz_str[0], gz_str[0])
                chi = chi_map.get(gz_str[1], gz_str[1])
                return f"{can} {chi}", can, chi

            nam, can_nam, chi_nam = translate_gz(ld.gz_year)
            thang, can_thang, chi_thang = translate_gz(ld.gz_month)
            ngay, can_ngay, chi_ngay = translate_gz(ld.gz_day)
            
            gio_chi_idx = ((dt.hour + 1) // 2) % 12
            chi_gio_str = list(chi_map.values())[gio_chi_idx]
            
            can_ngay_idx = list(can_map.values()).index(can_ngay)
            start_can_gio = (can_ngay_idx % 5) * 2
            can_gio_idx = (start_can_gio + gio_chi_idx) % 10
            can_gio_str = list(can_map.values())[can_gio_idx]
            gio = f"{can_gio_str} {chi_gio_str}"
        except Exception:
            nam, can_nam, chi_nam = "Ất Sửu", "Ất", "Sửu"
            thang, can_thang, chi_thang = "Bính Tuất", "Bính", "Tuất"
            ngay, can_ngay, chi_ngay = "Tân Mão", "Tân", "Mão"
            gio, can_gio_str, chi_gio_str = "Ất Mùi", "Ất", "Mùi"
            
        return {
            "nam": nam, "thang": thang, "ngay": ngay, "gio": gio,
            "can_nam": can_nam, "chi_nam": chi_nam,
            "can_thang": can_thang, "chi_thang": chi_thang,
            "can_ngay": can_ngay, "chi_ngay": chi_ngay,
            "can_gio": can_gio_str, "chi_gio": chi_gio_str
        }

    def calculate_tich_nhat(self, dt: datetime.datetime):
        # Thái Ất Kể Ngày: 
        # 1. Tìm Đông Chí (Term 23) của năm trước
        dong_chi_date = TermUtils.nth_term_day(dt.year - 1, 23)
        
        # 2. Tìm ngày Giáp Tý đầu tiên sau Đông Chí
        curr_date = dong_chi_date + timedelta(days=1)
        while True:
            ld = LunarDate.from_solar_date(curr_date.year, curr_date.month, curr_date.day)
            if ld.gz_day == '甲子':
                break
            curr_date += timedelta(days=1)
            
        # 3. Tính số ngày tích (Tích nhật)
        tich_nhat = (dt.date() - curr_date).days + 1
        return tich_nhat

    def calculate_nguyen_cuc(self, tich_nhat: int):
        if tich_nhat <= 0: return {"nguyen": 0, "cuc": 0, "khoi": "?"}
        nguyen = (tich_nhat // 72) + 1
        cuc = tich_nhat % 72
        if cuc == 0:
            cuc = 72; nguyen -= 1
        return {"nguyen": nguyen, "cuc": cuc, "khoi": "Dương"}

    def calculate_toan(self, stars_cung, dt: datetime.datetime=None):
        toan_data = stars_cung.get("_toan", {"chu": 0, "khach": 0, "dinh": 0})
        
        def format_toan(val):
            if val < 10:
                return f"{val} (Vô Thiên)"
            elif val % 10 == 0:
                return f"{val} (Vô Nhân)"
            elif val % 10 < 5:
                return f"{val} (Vô Địa)"
            return str(val)
        
        # --- TOÁN HÒA / KHÔNG HÒA ---
        # Chính cung (8 cung mang biệt số)
        chinh_cung = {"Kiền": 1, "Ngọ": 2, "Cấn": 3, "Mão": 4, "Dậu": 6, "Khôn": 7, "Tý": 8, "Tốn": 9}
        # Gián thần (12 chi xen kẽ)
        gian_than = ["Sửu", "Dần", "Thìn", "Tỵ", "Mùi", "Thân", "Tuất", "Hợi"]
        
        # Thái Ất: Cung Dương = biệt số 3,8,4,9 / Cung Âm = biệt số 2,7,6,1
        at_cung_duong = {3, 8, 4, 9}
        at_cung_am = {2, 7, 6, 1}
        
        # Ngoại lệ Nhị Mục
        nhi_muc_hoa_exception = {11, 13, 17, 19, 33, 37, 39}
        nhi_muc_bat_hoa_exception = {22, 24, 26, 28}
        
        def check_hoa_nhi_muc(star_pos, toan_val):
            """Xét Hòa cho Nhị Mục (Văn Xương / Thủy Kích)"""
            is_chan = (toan_val % 2 == 0)
            is_chinh = star_pos in chinh_cung  # Cung Dương
            is_gian = star_pos in gian_than    # Cung Âm
            
            # Ngoại lệ đặc biệt
            if toan_val in nhi_muc_hoa_exception and is_chinh:
                return "Hòa"  # Khí Hòa đặc biệt
            if toan_val in nhi_muc_bat_hoa_exception and is_gian:
                return "Không hòa"  # Khí Bất Hòa tuyệt đối
            
            # Tiêu chuẩn: Dương+Chẵn=Hòa, Âm+Lẻ=Hòa
            if (is_chinh and is_chan) or (is_gian and not is_chan):
                return "Hòa"
            return "Không hòa"
        
        def check_hoa_thai_at(cung_so, toan_val):
            """Xét Hòa cho Thái Ất (Ất Cả)"""
            is_chan = (toan_val % 2 == 0)
            is_duong = cung_so in at_cung_duong
            is_am = cung_so in at_cung_am
            
            # Dương+Chẵn=Hòa, Âm+Lẻ=Hòa
            if (is_duong and is_chan) or (is_am and not is_chan):
                return "Hòa"
            return "Không hòa"
        
        # Xét từng toán
        van_xuong_pos = toan_data.get("van_xuong_pos", "Dần")
        thuy_kich_pos = toan_data.get("thuy_kich_pos", "Kiền")
        thai_at_cung = toan_data.get("thai_at_cung", 9)
        
        hoa_chu = check_hoa_nhi_muc(van_xuong_pos, toan_data["chu"])
        hoa_khach = check_hoa_nhi_muc(thuy_kich_pos, toan_data["khach"])
        hoa_dinh = check_hoa_thai_at(thai_at_cung, toan_data["dinh"])
        
        chu = format_toan(toan_data["chu"])
        khach = format_toan(toan_data["khach"])
        dinh = format_toan(toan_data["dinh"])
        
        return {
            "chu": f"Toán chủ {chu}. {hoa_chu}",
            "khach": f"Toán khách {khach}. {hoa_khach}",
            "dinh": f"Toán định {dinh}. {hoa_dinh}"
        }

    def _get_chi_index(self, chi):
        chis = ["Tý", "Sửu", "Dần", "Mão", "Thìn", "Tỵ", "Ngọ", "Mùi", "Thân", "Dậu", "Tuất", "Hợi"]
        return chis.index(chi) if chi in chis else 0
        
    def _get_can_index(self, can):
        cans = ["Giáp", "Ất", "Bính", "Đinh", "Mậu", "Kỷ", "Canh", "Tân", "Nhâm", "Quý"]
        return cans.index(can) if can in cans else 0

    def calculate_nap_am_ngu_hanh(self, can, chi):
        # 1=Kim, 2=Thủy, 3=Hỏa, 4=Thổ, 5=Mộc
        can_idx = self._get_can_index(can)
        chi_idx = self._get_chi_index(chi)
        can_val = (can_idx // 2) + 1
        chi_val = (chi_idx // 2) % 3
        if chi_val == 0: chi_val = 0
        elif chi_val == 1: chi_val = 1
        else: chi_val = 2
        
        # Correct Chi Val map: Tý/Sửu=0, Dần/Mão=1, Thìn/Tỵ=2, Ngọ/Mùi=0, Thân/Dậu=1, Tuất/Hợi=2
        chi_val = (chi_idx // 2) % 3
        
        total = can_val + chi_val
        if total > 5: total -= 5
        return total

    def get_sinh_thanh(self, hanh_val):
        # 1=Kim(13), 2=Thủy(7), 3=Hỏa(9), 4=Thổ(15), 5=Mộc(11)
        return {1: 13, 2: 7, 3: 9, 4: 15, 5: 11}.get(hanh_val, 0)
        
    def get_can_hanh(self, can):
        # Mộc(5), Hỏa(3), Thổ(4), Kim(1), Thủy(2)
        idx = self._get_can_index(can)
        return [5, 5, 3, 3, 4, 4, 1, 1, 2, 2][idx]
        
    def get_chi_hanh(self, chi):
        # Tý(2), Sửu(4), Dần(5), Mão(5), Thìn(4), Tỵ(3), Ngọ(3), Mùi(4), Thân(1), Dậu(1), Tuất(4), Hợi(2)
        idx = self._get_chi_index(chi)
        return [2, 4, 5, 5, 4, 3, 3, 4, 1, 1, 4, 2][idx]

    def _calc_tru_value(self, bdtu, p):
        """Tính Số Thu Mầm cho 1 trụ (Can + Chi + Nạp Âm sinh thành)"""
        can = bdtu.get(f"can_{p}")
        chi = bdtu.get(f"chi_{p}")
        if not can or not chi: return 0
        can_st = self.get_sinh_thanh(self.get_can_hanh(can))
        chi_st = self.get_sinh_thanh(self.get_chi_hanh(chi))
        na_st = self.get_sinh_thanh(self.calculate_nap_am_ngu_hanh(can, chi))
        return can_st + chi_st + na_st

    def calculate_tong_so_thu_mam(self, bdtu):
        """Tổng cả 4 trụ (dùng cho Quẻ Dựng Nghiệp)"""
        return sum(self._calc_tru_value(bdtu, p) for p in ["nam", "thang", "ngay", "gio"])

    def calculate_tong_ngay_gio(self, bdtu):
        """Tổng chỉ trụ Ngày + trụ Giờ (dùng cho Hạn Chịu Khí)"""
        return sum(self._calc_tru_value(bdtu, p) for p in ["ngay", "gio"])

    def get_outer_rings(self, bdtu, tich_nhat, gioi_tinh="Nam"):
        # 1. Dương Cửu (Từ Can Ngày Sinh)
        can_ngay = bdtu.get("can_ngay", "Giáp")
        dc_khoi = {"Bính": (1, "Thân"), "Tân": (1, "Thân"),
                   "Mậu": (2, "Dần"), "Quý": (2, "Dần"),
                   "Đinh": (3, "Hợi"), "Nhâm": (3, "Hợi"),
                   "Ất": (4, "Tỵ"), "Canh": (4, "Tỵ"),
                   "Giáp": (5, "Ngọ"), "Kỷ": (5, "Ngọ")}
        dc_tuoi, dc_cung = dc_khoi.get(can_ngay, (5, "Ngọ"))
        
        # 2. Bách Lục — Luật thép: chỉ Ngày + Giờ + 55
        tong_ngay_gio = self.calculate_tong_ngay_gio(bdtu) + 55
        so_han = tong_ngay_gio % 60
        if so_han == 0: so_han = 60
        hoa_giap_idx = 0
        for i in range(60):
            if self.thien_can[i % 10] == bdtu.get("can_ngay", "Giáp") and self.dia_chi[i % 12] == bdtu.get("chi_ngay", "Tý"):
                hoa_giap_idx = i
                break
        chiu_khi_idx = (hoa_giap_idx - (so_han - 1)) % 60
        can_chiu_khi = self.thien_can[chiu_khi_idx % 10]
        bl_tuoi, bl_cung = dc_khoi.get(can_chiu_khi, (5, "Ngọ"))
        
        # 3. Lộc Mã (Từ Can Năm)
        can_nam = bdtu.get("can_nam", "Giáp")
        lm_khoi = {"Nhâm": (1, "Thân"), "Quý": (1, "Thân"),
                   "Bính": (2, "Dần"), "Đinh": (2, "Dần"),
                   "Giáp": (3, "Hợi"), "Ất": (3, "Hợi"),
                   "Canh": (4, "Tỵ"), "Tân": (4, "Tỵ"),
                   "Mậu": (5, "Ngọ"), "Kỷ": (5, "Ngọ")}
        lm_tuoi, lm_cung = lm_khoi.get(can_nam, (3, "Hợi"))
        
        chis = ["Tý", "Sửu", "Dần", "Mão", "Thìn", "Tỵ", "Ngọ", "Mùi", "Thân", "Dậu", "Tuất", "Hợi"]
        rings = {c: [] for c in chis}
        
        # Điền Dương Cửu (Thuận)
        curr_idx = chis.index(dc_cung)
        rings[chis[curr_idx]].append(f"1-{dc_tuoi} (Dương cửu)" if dc_tuoi > 1 else "1 (Dương cửu)")
        curr_idx = (curr_idx + 1) % 12
        for i in range(1, 8):
            s = dc_tuoi + 1 + (i-1)*10
            rings[chis[curr_idx]].append(f"{s}-{s+9}")
            curr_idx = (curr_idx + 1) % 12
            
        # Điền Bách Lục (Nam thuận +1, Nữ nghịch -1)
        bl_dir = 1 if gioi_tinh == "Nam" else -1
        curr_idx = chis.index(bl_cung)
        rings[chis[curr_idx]].append(f"1-{bl_tuoi} (Bách lục)" if bl_tuoi > 1 else "1 (Bách lục)")
        curr_idx = (curr_idx + bl_dir) % 12
        for i in range(1, 8):
            s = bl_tuoi + 1 + (i-1)*10
            rings[chis[curr_idx]].append(f"{s}-{s+9}")
            curr_idx = (curr_idx + bl_dir) % 12
            
        # Điền Lộc (Xuôi), Mã (Ngược)
        loc_idx = chis.index(lm_cung)
        ma_idx = chis.index(lm_cung)
        for i in range(8):
            s = lm_tuoi + i*10
            rings[chis[loc_idx]].append(f"Lộc {s}-{s+9}" if i == 0 else f"Lộc {s}")
            rings[chis[ma_idx]].append(f"Mã {s}-{s+9}" if i == 0 else f"Mã {s}")
            loc_idx = (loc_idx + 1) % 12
            ma_idx = (ma_idx - 1) % 12

        # --- ĐẠI HẠN THÁI ẤT LƯU NIÊN (Mỗi năm dời 1 cung) ---
        can_nam = bdtu.get("can_nam", "Giáp")
        can_thang = bdtu.get("can_thang", "Giáp")
        chi_thang = bdtu.get("chi_thang", "Tý")
        
        # Số năm đầu theo ngũ hành địa bàn của Chi Tháng (Thủy->1, Hỏa->2, Mộc->3, Kim->4, Thổ->5)
        chi_ngu_hanh_years = {
            "Tý": 1, "Hợi": 1,
            "Tỵ": 2, "Ngọ": 2,
            "Dần": 3, "Mão": 3,
            "Thân": 4, "Dậu": 4,
            "Thìn": 5, "Tuất": 5, "Sửu": 5, "Mùi": 5
        }
        start_years = chi_ngu_hanh_years.get(chi_thang, 5)
        
        # Chiều đi: Dương Nam / Âm Nữ đi Thuận (+1), Âm Nam / Dương Nữ đi Nghịch (-1)
        is_nam = (gioi_tinh == "Nam")
        is_duong_nam = can_nam in ["Giáp", "Bính", "Mậu", "Canh", "Nhâm"]
        direction = 1 if (is_duong_nam == is_nam) else -1
        
        # Cung khởi đầu: Chi Tháng sinh
        start_idx = chis.index(chi_thang)
        
        dai_han = {}
        for k in range(12):
            curr_idx = (start_idx + k * direction) % 12
            chi = chis[curr_idx]
            
            # Tính các tuổi đại hạn rơi vào cung này (lên đến 80 tuổi)
            ages = []
            if k == 0:
                if start_years > 1:
                    ages_str = f"1-{start_years}"
                else:
                    ages_str = "1"
                y_val = start_years + 12
                while y_val <= 80:
                    ages.append(y_val)
                    y_val += 12
            else:
                first_age = start_years + k
                y_val = first_age + 12
                while y_val <= 80:
                    ages.append(y_val)
                    y_val += 12
                ages_str = str(first_age)
                
            if ages:
                ages_suffix = ", ".join(map(str, ages))
                dai_han[chi] = f"ĐH {ages_str} ({ages_suffix})"
            else:
                dai_han[chi] = f"ĐH {ages_str}"

        def r_filter(chi, *kw_include):
            return "\n".join(x for x in rings[chi] if any(k in x for k in kw_include))
        def r_exclude(chi, *kw_exclude):
            return "\n".join(x for x in rings[chi] if not any(k in x for k in kw_exclude))

        return {
            "top": [
                [""] * 15,
                ["","","","","","", dai_han.get("Tỵ",""), dai_han.get("Ngọ",""), dai_han.get("Mùi",""), "","","","","",""],
                ["","","","","","", r_filter("Tỵ","Mã"), r_filter("Ngọ","Mã"), r_filter("Mùi","Mã"), "","","","","",""],
                ["","","","","","", r_filter("Tỵ","Lộc"), r_filter("Ngọ","Lộc"), r_filter("Mùi","Lộc"), "","","","","",""],
                ["","","","","","", r_exclude("Tỵ","Lộc","Mã"), r_exclude("Ngọ","Lộc","Mã"), r_exclude("Mùi","Lộc","Mã"), "","","","","",""]
            ],
            "left": [
                ["","","",""],
                [dai_han.get("Thìn",""), r_filter("Thìn","Lộc"), r_exclude("Thìn","Lộc","Mã"), ""],
                [dai_han.get("Mão",""), r_filter("Mão","Lộc"), r_exclude("Mão","Lộc","Mã"), ""],
                [dai_han.get("Dần",""), r_filter("Dần","Lộc"), r_exclude("Dần","Lộc","Mã"), ""],
                ["","","",""]
            ],
            "right": [
                ["","","",""],
                [r_exclude("Thân","Lộc","Mã"), "", r_filter("Thân","Mã"), dai_han.get("Thân","")],
                [r_exclude("Dậu","Lộc","Mã"), "", r_filter("Dậu","Mã"), dai_han.get("Dậu","")],
                [r_exclude("Tuất","Lộc","Mã"), "", r_filter("Tuất","Mã"), dai_han.get("Tuất","")],
                ["","","",""]
            ],
            "bottom": [
                ["","","","","","", r_exclude("Sửu","Lộc","Mã"), r_exclude("Tý","Lộc","Mã"), r_exclude("Hợi","Lộc","Mã"), "","","","","",""],
                ["","","","","","", dai_han.get("Sửu",""), dai_han.get("Tý",""), dai_han.get("Hợi",""), "","","","","",""],
                ["","","","","","", r_filter("Sửu","Lộc","Mã"), r_filter("Tý","Lộc","Mã"), r_filter("Hợi","Lộc","Mã"), "","","","","",""],
                [""] * 15
            ]
        }



    def calculate_que_dich(self, bdtu):
        # --- QUẺ DỰNG NGHIỆP: Dùng tổng CẢ 4 TRỤ + 55 ---
        tich_so_4tru = self.calculate_tong_so_thu_mam(bdtu) + 55
        que_idx = tich_so_4tru % 64
        if que_idx == 0: que_idx = 64
        que_goc = ThaiAtCalculator.HEXAGRAMS[que_idx]
        
        # --- HẠN CHỊU KHÍ: Chỉ dùng trụ NGÀY + trụ GIỜ + 55 ---
        tong_ngay_gio = self.calculate_tong_ngay_gio(bdtu) + 55
        so_han = tong_ngay_gio % 60
        if so_han == 0: so_han = 60
        
        # Tìm vị trí ngày sinh trên vòng 60 Hoa Giáp
        can_ngay = bdtu.get("can_ngay", "Giáp")
        chi_ngay = bdtu.get("chi_ngay", "Tý")
        c_idx = self._get_can_index(can_ngay)
        ch_idx = self._get_chi_index(chi_ngay)
        # Vị trí Can Chi trên vòng 60 Hoa Giáp (0-indexed)
        ngay_sinh_idx = (c_idx + (((c_idx - ch_idx) % 12) // 2) * 10) % 60
        # Sửa: Dùng công thức chuẩn tìm thứ tự trên 60 Giáp Tý
        # Can Chi -> Hoa Giáp index: (chi_idx * 5 + can_idx // 2) nếu cùng tính chẵn lẻ
        hoa_giap_idx = -1
        for i in range(60):
            if self.thien_can[i % 10] == can_ngay and self.dia_chi[i % 12] == chi_ngay:
                hoa_giap_idx = i
                break
        
        # Đếm ngược so_han ngày từ ngày sinh (ngày sinh tính là 1)
        chiu_khi_idx = (hoa_giap_idx - (so_han - 1)) % 60
        
        can_chiu_khi = self.thien_can[chiu_khi_idx % 10]
        chi_chiu_khi = self.dia_chi[chiu_khi_idx % 12]
        chi_val = (chiu_khi_idx % 12) + 1
        
        is_duong = can_chiu_khi in ["Giáp", "Bính", "Mậu", "Canh", "Nhâm"]
        target_bit = '1' if is_duong else '0'
        
        # Dương đi từ dưới lên trên (0 đến 5). Âm đi từ trên xuống dưới (5 đến 0).
        indices = list(range(6)) if is_duong else list(range(5, -1, -1))
        
        binary = que_goc["binary"]
        valid_indices = [i for i in indices if binary[i] == target_bit]
        
        if len(valid_indices) == 0:
            return {
                "que_goc": que_goc["name"], "hao_dong": "Không có Hào Động", "que_bien": que_goc["name"], "chi_tiet": f"Ngày chịu khí {can_chiu_khi} {chi_chiu_khi}"
            }
            
        hao_idx = valid_indices[(chi_val - 1) % len(valid_indices)]
        
        new_binary = list(binary)
        new_binary[hao_idx] = '1' if binary[hao_idx] == '0' else '0'
        new_binary = "".join(new_binary)
        
        que_bien_name = ""
        for k, v in ThaiAtCalculator.HEXAGRAMS.items():
            if v["binary"] == new_binary:
                que_bien_name = v["name"]
                break
                
        hao_names = ["Hào Sơ", "Hào 2", "Hào 3", "Hào 4", "Hào 5", "Hào Thượng"]
        hao_name_full = f"{hao_names[hao_idx]} {'Dương' if binary[hao_idx]=='1' else 'Âm'} động"
        
        return {
            "que_goc": f"Quẻ {que_idx}: {que_goc['name']}",
            "hao_dong": hao_name_full,
            "que_bien": f"Quẻ Biến: {que_bien_name}",
            "chi_tiet": f"(Tổng4Trụ {tich_so_4tru}, NgàyGiờ {tong_ngay_gio}, Hạn {so_han}, Khí {can_chiu_khi} {chi_chiu_khi})"
        }
        
    def build_5x5_grid(self, stars_cung, stars_chi, toan, bdtu, gioi_tinh="Nam"):
        grid = [[[] for _ in range(5)] for _ in range(5)]
        
        # Grid positions for the 16 Cung
        grid_positions = {
            "Tốn": (0, 0), "Tỵ": (0, 1), "Ngọ": (0, 2), "Mùi": (0, 3), "Khôn": (0, 4),
            "Thân": (1, 4), "Dậu": (2, 4), "Tuất": (3, 4), "Kiền": (4, 4),
            "Hợi": (4, 3), "Tý": (4, 2), "Sửu": (4, 1), "Cấn": (4, 0),
            "Dần": (3, 0), "Mão": (2, 0), "Thìn": (1, 0)
        }
        
        base_labels = {
            (0,0): ["9 TỐN mộc"], (0,1): ["TỴ hỏa"], (0,2): ["2.NGỌ hỏa"], (0,3): ["Vị thổ"], (0,4): ["7.KHÔN Thổ"],
            (1,0): ["THÌN thổ"], (1,4): ["Thân kim"],
            (2,0): ["MÃO mộc"], (2,4): ["6.Dậu kim"],
            (3,0): ["DẦN mộc"], (3,4): ["Tuất thổ"],
            (4,0): ["3 CẤN thổ"], (4,1): ["SỬU thổ"], (4,2): ["Tý thủy"], (4,3): ["HỢI thủy"], (4,4): ["1.Càn kim"]
        }
        for (r, c), lbls in base_labels.items():
            grid[r][c] = lbls.copy()
            
        chis = ["Tý", "Sửu", "Dần", "Mão", "Thìn", "Tỵ", "Ngọ", "Mùi", "Thân", "Dậu", "Tuất", "Hợi"]
        chi_nam = bdtu.get("chi_nam", "Tý")
        can_nam = bdtu.get("can_nam", "Giáp")
        chi_thang = bdtu.get("chi_thang", "Tý")
        chi_ngay = bdtu.get("chi_ngay", "Tý")
        chi_gio = bdtu.get("chi_gio", "Tý")
        
        is_nam = (gioi_tinh == "Nam")
        is_duong_nam = can_nam in ["Giáp", "Bính", "Mậu", "Canh", "Nhâm"]
        direction = 1 if (is_duong_nam == is_nam) else -1
        
        y_idx = chis.index(chi_nam)
        m_idx = chis.index(chi_thang)
        d_idx = chis.index(chi_ngay)
        h_idx = chis.index(chi_gio)
        
        steps_menh = (h_idx - m_idx) % 12
        if steps_menh < 0: steps_menh += 12
        menh_idx = (y_idx + steps_menh * direction) % 12
        
        # Thân: An trực tiếp tại cung mang tên Chi Ngày sinh (Thái Ất Nhân Mệnh)
        than_idx = d_idx  # Chi Ngày sinh
        
        cung_names = ["MỆNH", "HUYNH", "THÊ" if is_nam else "PHU", "TỬ", "TÀI", "ĐIỀN", "QUAN", "NÔ", "TẬT", "PHÚC ĐỨC", "TƯỚNG", "PHỤ MẪU"]
        
        for i, name in enumerate(cung_names):
            c_idx = (menh_idx + i * direction) % 12
            chi_name = chis[c_idx]
            coord = grid_positions[chi_name]
            
            lbl = name
            if c_idx == than_idx:
                lbl += " (THÂN)"
                
            grid[coord[0]][coord[1]].insert(0, lbl)
        
        # If Thân is not one of the 12 Cung (it always is, but just in case)
        # Actually it's guaranteed because 12 Cung covers all 12 Chi.
        
        for star, chi in stars_chi.items():
            r, c = ThaiAtStarsMap.GRID_MAPPING[chi]
            grid[r][c].append(star)
            
        for star, cung in stars_cung.items():
            if star == "_toan":
                continue
            if cung in ThaiAtStarsMap.CUNG_GOC_TO_NAME:
                name = ThaiAtStarsMap.CUNG_GOC_TO_NAME[cung]
                r, c = ThaiAtStarsMap.GRID_MAPPING[name]
                grid[r][c].append(star)
            else:
                chi_list = ThaiAtStarsMap.CUU_CUNG_TO_12_CHI.get(cung, [])
                if chi_list:
                    r, c = ThaiAtStarsMap.GRID_MAPPING[chi_list[0]]
                    grid[r][c].append(star)
                    
        # Inner cells
        grid[1][1] = ["Đại Thần"]
        grid[1][2] = ["Đại uy"]
        grid[1][3] = ["Thiên Đạo"]
        grid[2][1] = ["Cao Tùng"]
        grid[2][3] = ["Vũ Đức"]
        grid[3][1] = ["Lã Thân"]
        grid[3][2] = ["Địa chủ", "Khí rời"]
        grid[3][3] = ["Âm chủ"]
        
        loc_chu_map = {
            "Giáp": "Quân Cơ", "Kỷ": "Quân Cơ",
            "Ất": "Đại Tướng Chủ", "Canh": "Đại Tướng Chủ",
            "Bính": "Đại Tướng Khách", "Tân": "Đại Tướng Khách",
            "Đinh": "Thái Ất", "Nhâm": "Thái Ất",
            "Mậu": "Thủy Kích", "Quý": "Thủy Kích"
        }
        loc_chu_star = loc_chu_map.get(can_nam, "Đại Tướng Chủ")
        
        trung_cung_data = [
            f"Sao Lộc Chủ: {loc_chu_star}", toan['chu'], toan['khach'], toan['dinh']
        ]
        
        # Calculate Quẻ Dịch
        try:
            que_data = self.calculate_que_dich(bdtu)
            trung_cung_data.extend([
                que_data["que_goc"],
                que_data["hao_dong"],
                que_data["que_bien"],
                que_data["chi_tiet"]
            ])
        except Exception as e:
            trung_cung_data.append(f"Quẻ: (Lỗi tính toán {e})")
            
        grid[2][2] = trung_cung_data + grid[2][2]
        
        return grid

    def calculate_hao_dong_luu_nien(self, q_ln, target_year):
        # 1. Can Chi of target_year
        offset = (target_year - 1984) % 60
        can_idx = (offset) % 10
        chi_idx = (offset) % 12
        
        can_name = self.thien_can[can_idx]
        chi_name = self.dia_chi[chi_idx]
        
        # 2. Year type: Dương or Âm
        is_duong_year = can_name in ["Giáp", "Bính", "Mậu", "Canh", "Nhâm"]
        
        # 3. Steps = Chi index (1-indexed)
        steps = chi_idx + 1
        
        # 4. Hexagram binary
        hex_info = self.HEXAGRAMS.get(q_ln)
        if not hex_info:
            return 1, "Hào Sơ", "Dương"
            
        binary = hex_info["binary"] # e.g. "110000" (index 0 is Hào 1)
        
        # 5. Filter available Hào
        target_bit = '1' if is_duong_year else '0'
        avail_hao = [i + 1 for i in range(6) if binary[i] == target_bit]
        
        if not avail_hao:
            # Fallback to all 6 Hào
            hao_num = (steps - 1) % 6 + 1
            bit_val = binary[hao_num - 1]
            hao_type = "Dương" if bit_val == '1' else "Âm"
        else:
            hao_num = avail_hao[(steps - 1) % len(avail_hao)]
            bit_val = binary[hao_num - 1]
            hao_type = "Dương" if bit_val == '1' else "Âm"
            
        hao_names = ["Hào Sơ", "Hào 2", "Hào 3", "Hào 4", "Hào 5", "Hào Thượng"]
        hao_name_str = f"{hao_names[hao_num - 1]} {hao_type} động"
        
        return hao_num, hao_name_str, f"{can_name} {chi_name}"

    def get_macro_monthly_cung(self, target_year):
        """
        Calculates the 4 macro monthly cụms for a given year based on the 2-year cyclic Thai At rule.
        Returns a dict mapping each Cụm name to a tuple: (Cung number, Cung Name, Element, Month list)
        """
        cungs = [1, 2, 3, 4, 6, 7, 8, 9]
        cung_elements = {
            1: ("Kiền", "Kim"),
            2: ("Ly", "Hỏa"),
            3: ("Cấn", "Thổ"),
            4: ("Chấn", "Mộc"),
            6: ("Đoài", "Kim"),
            7: ("Khôn", "Thổ"),
            8: ("Khảm", "Thủy"),
            9: ("Tốn", "Mộc")
        }
        
        if target_year % 2 == 0:
            indices = [1, 2, 3, 4]
        else:
            indices = [5, 6, 7, 0]
            
        c1_cung = cungs[indices[0]]
        c2_cung = cungs[indices[1]]
        c3_cung = cungs[indices[2]]
        c4_cung = cungs[indices[3]]
        
        c1_name, c1_elem = cung_elements[c1_cung]
        c2_name, c2_elem = cung_elements[c2_cung]
        c3_name, c3_elem = cung_elements[c3_cung]
        c4_name, c4_elem = cung_elements[c4_cung]
        
        return {
            "Cụm1_T234": (c1_cung, c1_name, c1_elem, [3, 4, 5]),
            "Cụm2_T567": (c2_cung, c2_name, c2_elem, [6, 7, 8]),
            "Cụm3_T8910": (c3_cung, c3_name, c3_elem, [9, 10, 11]),
            "Cụm4_T11121": (c4_cung, c4_name, c4_elem, [12, 1, 2])
        }

    def calculate_all(self, dt: datetime.datetime, gioi_tinh: str = "Nam"):
        bat_tu = self.calculate_bat_tu(dt)
        tich_nhat = self.calculate_tich_nhat(dt)
        nguyen_cuc = self.calculate_nguyen_cuc(tich_nhat)
        
        stars_cung, stars_chi = ThaiAtStarsMap.map_stars(tich_nhat, nguyen_cuc["khoi"] == "Dương", bat_tu)
        toan = self.calculate_toan(stars_cung, dt)
        grid_5x5 = self.build_5x5_grid(stars_cung, stars_chi, toan, bat_tu, gioi_tinh)
        outer_rings = self.get_outer_rings(bat_tu, tich_nhat, gioi_tinh)
        
        return {
            "datetime": dt.strftime("%Y-%m-%d %H:%M:%S"),
            "bat_tu": bat_tu, "tich_nhat": tich_nhat, "nguyen_cuc": nguyen_cuc,
            "toan": toan, "grid_5x5": grid_5x5, "outer": outer_rings
        }

    def export_excel(self, data: dict, output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "Bàn Cờ Thái Ất"
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        start_row = 1
        grid = data["grid_5x5"]
        outer = data["outer"]
        
        for r_idx, row_data in enumerate(outer["top"]):
            for c_idx, val in enumerate(row_data):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=val)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        
        start_row += len(outer["top"])
        for r in range(5):
            left = outer["left"][r]
            right = outer["right"][r]
            for c in range(4):
                cell = ws.cell(row=start_row + r, column=c + 2, value=left[c])
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            for c in range(5):
                cell_val = "\n".join(grid[r][c])
                cell = ws.cell(row=start_row + r, column=c + 6, value=cell_val)
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            for c in range(4):
                cell = ws.cell(row=start_row + r, column=c + 11, value=right[c])
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                
        start_row += 5
        for r_idx, row_data in enumerate(outer["bottom"]):
            for c_idx, val in enumerate(row_data):
                cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=val)
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                
        # Resize columns and rows
        for col in range(1, 16): ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'+chr(64+col-26)].width = 15
        for r in range(1, 20): ws.row_dimensions[r].height = 40
        ws.row_dimensions[8].height = 100 # Center row height
        
        # Thêm sheet Luận Giải
        if data["tich_nhat"] == 268:
            try:
                import os
                if os.path.exists("luan_giai_at_suu.md"):
                    with open("luan_giai_at_suu.md", "r", encoding="utf-8") as f:
                        ws2 = wb.create_sheet(title="Luận Giải Chi Tiết")
                        for i, line in enumerate(f.readlines()):
                            cell = ws2.cell(row=i+1, column=1, value=line.strip())
                            cell.alignment = Alignment(wrap_text=True)
                        ws2.column_dimensions['A'].width = 120
            except: pass
            
        wb.save(output_path)
        return output_path
